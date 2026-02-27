from pathlib import Path

from nigel.importer import excel_serial_to_date
from nigel.models import ParsedRow


def detect_payroll(file_path: Path) -> bool:
    """Return True if file looks like a Gusto payroll XLSX."""
    if file_path.suffix.lower() != ".xlsx":
        return False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        has_payrolls = "payrolls" in wb.sheetnames
        wb.close()
        return has_payrolls
    except Exception:
        return False


def parse_payroll(file_path: Path) -> list[ParsedRow]:
    """Parse a Gusto payroll XLSX. Aggregates per pay period."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)

    payrolls_ws = wb["payrolls"]
    payroll_rows = list(payrolls_ws.iter_rows(min_row=2, values_only=True))
    wages_by_date: dict[str, float] = {}
    for row in payroll_rows:
        if row[3] is None or row[7] is None:
            continue
        check_date = excel_serial_to_date(row[3]) if isinstance(row[3], (int, float)) else str(row[3])
        gross = float(row[7])
        wages_by_date[check_date] = wages_by_date.get(check_date, 0.0) + gross

    taxes_ws = wb["taxes"]
    tax_rows = list(taxes_ws.iter_rows(min_row=2, values_only=True))
    employer_taxes_by_date: dict[str, float] = {}
    for row in tax_rows:
        if row[6] != "Employer" or row[3] is None or row[7] is None:
            continue
        check_date = excel_serial_to_date(row[3]) if isinstance(row[3], (int, float)) else str(row[3])
        amount = float(row[7])
        employer_taxes_by_date[check_date] = employer_taxes_by_date.get(check_date, 0.0) + amount

    wb.close()

    result: list[ParsedRow] = []
    for date, total in sorted(wages_by_date.items()):
        result.append(ParsedRow(
            date=date,
            description=f"Payroll — Wages ({date})",
            amount=-abs(total),
        ))
    for date, total in sorted(employer_taxes_by_date.items()):
        result.append(ParsedRow(
            date=date,
            description=f"Payroll — Employer Taxes ({date})",
            amount=-abs(total),
        ))

    return result
