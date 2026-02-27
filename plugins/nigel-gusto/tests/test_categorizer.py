import openpyxl

from nigel_gusto.payroll import parse_payroll
from nigel_gusto.categorizer import auto_categorize_payroll


def _create_gusto_fixture(path):
    """Create a minimal Gusto XLSX fixture."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "payrolls"
    ws.append(["Id", "Employee id", "Employee name", "Check date", "Payment period start",
               "Payment period end", "Payment method", "Gross pay", "Net pay"])
    ws.append(["payrolls_aaa", "emp1", "Doe, Jane", 45667, 45649, 45662, "Direct Deposit", 4000.00, 3000.00])
    ws.append(["payrolls_aaa", "emp2", "Doe, John", 45667, 45649, 45662, "Direct Deposit", 3500.00, 2600.00])

    ws2 = wb.create_sheet("taxes")
    ws2.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Tax", "Type", "Amount", "Subject wage", "Gross subject wage"])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 248.00, 4000.00, 4000.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 217.00, 3500.00, 3500.00])

    ws3 = wb.create_sheet("deductions")
    ws3.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Name", "Type", "Employee deduction", "Employer contribution"])
    wb.save(path)


def test_auto_categorize_payroll(db, tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)

    db.execute("INSERT INTO accounts (name, account_type) VALUES (?, ?)", ("Gusto", "payroll"))
    db.commit()

    rows = parse_payroll(fixture)

    for row in rows:
        db.execute(
            "INSERT INTO transactions (account_id, date, description, amount, is_flagged, flag_reason) "
            "VALUES (?, ?, ?, ?, 1, 'No matching rule')",
            (1, row.date, row.description, row.amount),
        )
    db.commit()

    auto_categorize_payroll(db, 1, rows)

    wages_txn = db.execute(
        "SELECT t.*, c.name as cat_name FROM transactions t "
        "LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.description LIKE '%Wages%'"
    ).fetchone()
    assert wages_txn["cat_name"] == "Payroll — Wages"
    assert wages_txn["is_flagged"] == 0

    tax_txn = db.execute(
        "SELECT t.*, c.name as cat_name FROM transactions t "
        "LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.description LIKE '%Taxes%'"
    ).fetchone()
    assert tax_txn["cat_name"] == "Payroll — Taxes"
    assert tax_txn["is_flagged"] == 0
