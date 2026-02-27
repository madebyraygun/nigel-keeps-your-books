import openpyxl

from nigel_gusto.payroll import parse_payroll, detect_payroll


def _create_gusto_fixture(path):
    """Create a minimal Gusto XLSX fixture programmatically."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "payrolls"
    ws.append(["Id", "Employee id", "Employee name", "Check date", "Payment period start",
               "Payment period end", "Payment method", "Gross pay", "Net pay"])
    ws.append(["payrolls_aaa", "emp1", "Doe, Jane", 45667, 45649, 45662, "Direct Deposit", 4000.00, 3000.00])
    ws.append(["payrolls_aaa", "emp2", "Doe, John", 45667, 45649, 45662, "Direct Deposit", 3500.00, 2600.00])
    ws.append(["payrolls_bbb", "emp1", "Doe, Jane", 45681, 45663, 45676, "Direct Deposit", 4000.00, 3000.00])
    ws.append(["payrolls_bbb", "emp2", "Doe, John", 45681, 45663, 45676, "Direct Deposit", 3500.00, 2600.00])

    ws2 = wb.create_sheet("taxes")
    ws2.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Tax", "Type", "Amount", "Subject wage", "Gross subject wage"])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 248.00, 4000.00, 4000.00])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Medicare", "Employer", 58.00, 4000.00, 4000.00])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "FUTA", "Employer", 24.00, 4000.00, 4000.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 217.00, 3500.00, 3500.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Medicare", "Employer", 50.75, 3500.00, 3500.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "FUTA", "Employer", 21.00, 3500.00, 3500.00])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Federal Income Tax", "Employee", 500.00, 4000.00, 4000.00])

    ws3 = wb.create_sheet("deductions")
    ws3.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Name", "Type", "Employee deduction", "Employer contribution"])
    wb.save(path)


def test_detect_payroll(tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)
    assert detect_payroll(fixture) is True

    csv_file = tmp_path / "not_gusto.csv"
    csv_file.write_text("Date,Description,Amount\n")
    assert detect_payroll(csv_file) is False


def test_parse_payroll(tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)

    rows = parse_payroll(fixture)
    wages_rows = [r for r in rows if "Wages" in r.description]
    tax_rows = [r for r in rows if "Taxes" in r.description]

    assert len(wages_rows) == 2
    assert wages_rows[0].amount == -7500.00
    assert tax_rows[0].amount == -618.75
