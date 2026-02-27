import csv
import hashlib
import sqlite3
from datetime import datetime
from pathlib import Path

from nigel.models import ImporterInfo, ParsedRow
from nigel.registry import registry


def _compute_checksum(file_path: Path) -> str:
    return hashlib.sha256(file_path.read_bytes()).hexdigest()


def _parse_amount(raw: str) -> float:
    """Strip commas and quotes, return float."""
    return float(raw.replace(",", "").replace('"', "").strip())


def _parse_date_mdy(raw: str) -> str:
    """Convert MM/DD/YYYY to ISO 8601 YYYY-MM-DD."""
    return datetime.strptime(raw.strip(), "%m/%d/%Y").strftime("%Y-%m-%d")


def parse_bofa_checking(file_path: Path) -> list[ParsedRow]:
    """Parse a BofA checking CSV, skipping the preamble rows."""
    rows: list[ParsedRow] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        # Find the header row: "Date,Description,Amount,Running Bal."
        for line in reader:
            if len(line) >= 4 and line[0].strip() == "Date" and "Description" in line[1]:
                break
        # Parse data rows
        for line in reader:
            if len(line) < 3 or not line[0].strip():
                continue
            try:
                date = _parse_date_mdy(line[0])
            except ValueError:
                continue
            description = line[1].strip()
            if not description or "Beginning balance" in description:
                continue
            amount = _parse_amount(line[2])
            rows.append(ParsedRow(date=date, description=description, amount=amount))
    return rows


def parse_bofa_credit_card(file_path: Path) -> list[ParsedRow]:
    """Parse a BofA credit card CSV. Amounts are always positive; use Transaction Type for sign."""
    rows: list[ParsedRow] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        # Find header row containing "Posting Date"
        for line in reader:
            if any("Posting Date" in cell for cell in line):
                break
        for line in reader:
            if len(line) < 10 or not line[2].strip():
                continue
            try:
                date = _parse_date_mdy(line[3])  # Trans. Date
            except ValueError:
                continue
            description = line[5].strip()
            amount = _parse_amount(line[6])
            txn_type = line[9].strip()
            # D = charge (expense), C = credit/payment
            if txn_type == "D":
                amount = -abs(amount)
            else:
                amount = abs(amount)
            rows.append(ParsedRow(date=date, description=description, amount=amount))
    return rows


def parse_bofa_line_of_credit(file_path: Path) -> list[ParsedRow]:
    """Parse a BofA line of credit CSV. Same columns as credit card but amounts are pre-signed."""
    rows: list[ParsedRow] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for line in reader:
            if any("Posting Date" in cell for cell in line):
                break
        for line in reader:
            if len(line) < 10 or not line[2].strip():
                continue
            try:
                date = _parse_date_mdy(line[3])  # Trans. Date
            except ValueError:
                continue
            description = line[5].strip()
            amount = _parse_amount(line[6])
            # Line of credit: positive = charge/expense (negate), negative = payment (keep)
            amount = -amount
            rows.append(ParsedRow(date=date, description=description, amount=amount))
    return rows


def _excel_serial_to_date(serial: int | float) -> str:
    """Convert Excel serial date number to ISO 8601 string."""
    from datetime import timedelta
    base = datetime(1899, 12, 30)  # Excel epoch (accounting for the 1900 leap year bug)
    return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")


def parse_gusto_payroll(file_path: Path) -> list[ParsedRow]:
    """Parse a Gusto payroll XLSX. Aggregates per pay period, discards employee detail."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Aggregate gross pay per check date from payrolls sheet
    payrolls_ws = wb["payrolls"]
    payroll_rows = list(payrolls_ws.iter_rows(min_row=2, values_only=True))
    # Group by check date (column index 3)
    wages_by_date: dict[str, float] = {}
    for row in payroll_rows:
        if row[3] is None or row[7] is None:
            continue
        check_date = _excel_serial_to_date(row[3]) if isinstance(row[3], (int, float)) else str(row[3])
        gross = float(row[7])
        wages_by_date[check_date] = wages_by_date.get(check_date, 0.0) + gross

    # Aggregate employer taxes per check date from taxes sheet
    taxes_ws = wb["taxes"]
    tax_rows = list(taxes_ws.iter_rows(min_row=2, values_only=True))
    employer_taxes_by_date: dict[str, float] = {}
    for row in tax_rows:
        if row[6] != "Employer" or row[3] is None or row[7] is None:
            continue
        check_date = _excel_serial_to_date(row[3]) if isinstance(row[3], (int, float)) else str(row[3])
        amount = float(row[7])
        employer_taxes_by_date[check_date] = employer_taxes_by_date.get(check_date, 0.0) + amount

    wb.close()

    # Build parsed rows — expenses are negative
    result: list[ParsedRow] = []
    for date, total in sorted(wages_by_date.items()):
        result.append(ParsedRow(
            date=date,
            description=f"Payroll — Wages ({date})",
            amount=-abs(total),
        ))
    for date, total in sorted(employer_taxes_by_date.items()):
        result.append(ParsedRow(
            date=date,
            description=f"Payroll — Employer Taxes ({date})",
            amount=-abs(total),
        ))

    return result


registry.register(ImporterInfo(
    key="bofa_checking", name="Bank of America Checking",
    account_types=["checking"], file_extensions=[".csv"],
    parse=parse_bofa_checking,
))
registry.register(ImporterInfo(
    key="bofa_credit_card", name="Bank of America Credit Card",
    account_types=["credit_card"], file_extensions=[".csv"],
    parse=parse_bofa_credit_card,
))
registry.register(ImporterInfo(
    key="bofa_line_of_credit", name="Bank of America Line of Credit",
    account_types=["line_of_credit"], file_extensions=[".csv"],
    parse=parse_bofa_line_of_credit,
))
registry.register(ImporterInfo(
    key="gusto_payroll", name="Gusto Payroll",
    account_types=["payroll"], file_extensions=[".xlsx"],
    parse=parse_gusto_payroll,
))


def _is_duplicate_row(conn: sqlite3.Connection, account_id: int, row: ParsedRow) -> bool:
    cursor = conn.execute(
        "SELECT 1 FROM transactions WHERE account_id = ? AND date = ? AND amount = ? AND description = ?",
        (account_id, row.date, row.amount, row.description),
    )
    return cursor.fetchone() is not None


def import_file(
    conn: sqlite3.Connection,
    file_path: Path,
    account_name: str,
) -> dict:
    """Import a CSV file into the database. Returns counts of imported/skipped."""
    # Look up account
    cursor = conn.execute(
        "SELECT id, account_type FROM accounts WHERE name = ?", (account_name,)
    )
    account = cursor.fetchone()
    if account is None:
        raise ValueError(f"Unknown account: {account_name}")

    account_id = account["id"]
    account_type = account["account_type"]

    # Check file-level duplicate
    checksum = _compute_checksum(file_path)
    cursor = conn.execute(
        "SELECT 1 FROM imports WHERE checksum = ? AND account_id = ?",
        (checksum, account_id),
    )
    if cursor.fetchone() is not None:
        return {"imported": 0, "skipped": 0, "duplicate_file": True}

    # Parse
    importer = registry.get_for_account_type(account_type)
    if importer is None:
        raise ValueError(f"No importer for account type: {account_type}")
    parsed_rows = importer.parse(file_path)

    # For payroll imports, look up categories for auto-assignment
    payroll_categories = {}
    if account_type == "payroll":
        for cat_name in ("Payroll — Wages", "Payroll — Taxes", "Payroll — Benefits"):
            row_cat = conn.execute("SELECT id FROM categories WHERE name = ?", (cat_name,)).fetchone()
            if row_cat:
                payroll_categories[cat_name] = row_cat["id"]

    # Insert
    imported = 0
    skipped = 0
    for row in parsed_rows:
        if _is_duplicate_row(conn, account_id, row):
            skipped += 1
            continue

        # Auto-categorize payroll transactions
        category_id = None
        is_flagged = 1
        flag_reason = "No matching rule"
        if account_type == "payroll":
            if "Wages" in row.description:
                category_id = payroll_categories.get("Payroll — Wages")
            elif "Taxes" in row.description:
                category_id = payroll_categories.get("Payroll — Taxes")
            elif "Benefits" in row.description:
                category_id = payroll_categories.get("Payroll — Benefits")
            if category_id:
                is_flagged = 0
                flag_reason = None

        conn.execute(
            "INSERT INTO transactions (account_id, date, description, amount, category_id, is_flagged, flag_reason) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (account_id, row.date, row.description, row.amount, category_id, is_flagged, flag_reason),
        )
        imported += 1

    # Record import batch
    dates = [r.date for r in parsed_rows]
    conn.execute(
        "INSERT INTO imports (filename, account_id, record_count, date_range_start, date_range_end, checksum) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            file_path.name,
            account_id,
            imported,
            min(dates) if dates else None,
            max(dates) if dates else None,
            checksum,
        ),
    )
    conn.commit()

    return {"imported": imported, "skipped": skipped, "duplicate_file": False}
