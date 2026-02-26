from pathlib import Path

from bookkeeper.importer import (
    parse_bofa_checking, parse_bofa_credit_card, parse_bofa_line_of_credit,
    parse_gusto_payroll, import_file,
)
from bookkeeper.db import init_db, get_connection

FIXTURES = Path(__file__).parent / "fixtures"


def test_parse_bofa_checking():
    rows = parse_bofa_checking(FIXTURES / "bofa_checking_sample.csv")
    assert len(rows) == 5
    # First transaction
    assert rows[0].date == "2024-12-02"
    assert rows[0].amount == -2500.00
    assert "ACME CORP" in rows[0].description
    # Positive amount (deposit)
    assert rows[3].date == "2025-01-09"
    assert rows[3].amount == 5000.00


def test_parse_bofa_checking_skips_preamble():
    rows = parse_bofa_checking(FIXTURES / "bofa_checking_sample.csv")
    for row in rows:
        assert "Beginning balance" not in row.description
        assert "Summary" not in row.description


def test_import_file_inserts_transactions(db):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA Checking", "checking"),
    )
    db.commit()

    result = import_file(db, FIXTURES / "bofa_checking_sample.csv", "BofA Checking")
    assert result["imported"] == 5
    assert result["skipped"] == 0

    cursor = db.execute("SELECT count(*) FROM transactions")
    assert cursor.fetchone()[0] == 5


def test_import_file_detects_file_duplicate(db):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA Checking", "checking"),
    )
    db.commit()

    import_file(db, FIXTURES / "bofa_checking_sample.csv", "BofA Checking")
    result = import_file(db, FIXTURES / "bofa_checking_sample.csv", "BofA Checking")
    assert result["duplicate_file"] is True


def test_import_file_detects_row_duplicates(db, tmp_path):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA Checking", "checking"),
    )
    db.commit()

    import_file(db, FIXTURES / "bofa_checking_sample.csv", "BofA Checking")

    # Create a different file with overlapping rows + one new row
    overlap_file = tmp_path / "overlap.csv"
    overlap_file.write_text(
        'Date,Description,Amount,Running Bal.\n'
        '12/02/2024,"ACME CORP DES:PAYMENT ID:12345 INDN:Raygun Design, LLC CO ID:XXXXX42850 CCD","-2,500.00","20,048.05"\n'
        '02/01/2025,"NEW VENDOR PAYMENT","-100.00","24,543.33"\n'
    )
    result = import_file(db, overlap_file, "BofA Checking")
    assert result["imported"] == 1
    assert result["skipped"] == 1

    cursor = db.execute("SELECT count(*) FROM transactions")
    assert cursor.fetchone()[0] == 6


def test_import_file_records_import_batch(db):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA Checking", "checking"),
    )
    db.commit()

    import_file(db, FIXTURES / "bofa_checking_sample.csv", "BofA Checking")

    cursor = db.execute("SELECT * FROM imports")
    record = cursor.fetchone()
    assert record["filename"] == "bofa_checking_sample.csv"
    assert record["record_count"] == 5
    assert record["checksum"] is not None


def test_parse_bofa_credit_card():
    rows = parse_bofa_credit_card(FIXTURES / "bofa_credit_card_sample.csv")
    assert len(rows) == 3
    # D type = charge, should be negative
    assert rows[0].amount == -54.43
    assert rows[0].date == "2025-03-10"  # Uses Trans. Date, not Posting Date
    assert "ADOBE" in rows[0].description
    # C type = payment, should be positive
    assert rows[2].amount == 500.00


def test_parse_bofa_line_of_credit():
    rows = parse_bofa_line_of_credit(FIXTURES / "bofa_loc_sample.csv")
    assert len(rows) == 3
    # Positive in CSV = charge → negate to -110.97
    assert rows[0].amount == -110.97
    assert "FINANCE CHARGE" in rows[0].description
    # Negative in CSV = payment → negate to +500.00
    assert rows[1].amount == 500.00


def test_import_credit_card(db):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA CC", "credit_card"),
    )
    db.commit()
    result = import_file(db, FIXTURES / "bofa_credit_card_sample.csv", "BofA CC")
    assert result["imported"] == 3


def test_import_line_of_credit(db):
    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("BofA LOC", "line_of_credit"),
    )
    db.commit()
    result = import_file(db, FIXTURES / "bofa_loc_sample.csv", "BofA LOC")
    assert result["imported"] == 3


import openpyxl


def _create_gusto_fixture(path):
    """Create a minimal Gusto XLSX fixture programmatically."""
    wb = openpyxl.Workbook()

    # payrolls sheet
    ws = wb.active
    ws.title = "payrolls"
    ws.append(["Id", "Employee id", "Employee name", "Check date", "Payment period start",
               "Payment period end", "Payment method", "Gross pay", "Net pay"])
    # Check date as Excel serial: 2025-01-10 = 45667
    ws.append(["payrolls_aaa", "emp1", "Doe, Jane", 45667, 45649, 45662, "Direct Deposit", 4000.00, 3000.00])
    ws.append(["payrolls_aaa", "emp2", "Doe, John", 45667, 45649, 45662, "Direct Deposit", 3500.00, 2600.00])
    # Second pay period: 2025-01-24 = 45681
    ws.append(["payrolls_bbb", "emp1", "Doe, Jane", 45681, 45663, 45676, "Direct Deposit", 4000.00, 3000.00])
    ws.append(["payrolls_bbb", "emp2", "Doe, John", 45681, 45663, 45676, "Direct Deposit", 3500.00, 2600.00])

    # taxes sheet
    ws2 = wb.create_sheet("taxes")
    ws2.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Tax", "Type", "Amount", "Subject wage", "Gross subject wage"])
    # Employer taxes for payrolls_aaa
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 248.00, 4000.00, 4000.00])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Medicare", "Employer", 58.00, 4000.00, 4000.00])
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "FUTA", "Employer", 24.00, 4000.00, 4000.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Social Security", "Employer", 217.00, 3500.00, 3500.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Medicare", "Employer", 50.75, 3500.00, 3500.00])
    ws2.append(["emp2", "Doe, John", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "FUTA", "Employer", 21.00, 3500.00, 3500.00])
    # Employee taxes (should be ignored for our purposes)
    ws2.append(["emp1", "Doe, Jane", "payrolls_aaa", 45667, "2024-12-23 - 2025-01-05",
                "Federal Income Tax", "Employee", 500.00, 4000.00, 4000.00])

    # deductions sheet (header only per real data)
    ws3 = wb.create_sheet("deductions")
    ws3.append(["Employee id", "Employee name", "Payroll id", "Payroll check date",
                "Payroll payment period", "Name", "Type", "Employee deduction", "Employer contribution"])

    wb.save(path)


def test_parse_gusto_payroll(tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)

    rows = parse_gusto_payroll(fixture)
    # Should produce 2 rows per pay period: wages + employer taxes
    # Pay period 1: wages=7500, employer taxes=618.75
    # Pay period 2: wages=7500, employer taxes (not in fixture, so just wages)
    wages_rows = [r for r in rows if "Wages" in r.description]
    tax_rows = [r for r in rows if "Taxes" in r.description]

    assert len(wages_rows) == 2
    assert wages_rows[0].amount == -7500.00
    assert tax_rows[0].amount == -618.75


def test_import_gusto_payroll(db, tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)

    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("Gusto", "payroll"),
    )
    db.commit()

    result = import_file(db, fixture, "Gusto")
    assert result["imported"] >= 2  # At least wages entries


def test_gusto_import_auto_categorizes(db, tmp_path):
    fixture = tmp_path / "gusto_sample.xlsx"
    _create_gusto_fixture(fixture)

    db.execute(
        "INSERT INTO accounts (name, account_type) VALUES (?, ?)",
        ("Gusto", "payroll"),
    )
    db.commit()

    import_file(db, fixture, "Gusto")

    # Wages should be auto-categorized
    wages_txn = db.execute(
        "SELECT t.*, c.name as cat_name FROM transactions t "
        "LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.description LIKE '%Wages%'"
    ).fetchone()
    assert wages_txn["cat_name"] == "Payroll — Wages"
    assert wages_txn["is_flagged"] == 0

    # Employer taxes should be auto-categorized
    tax_txn = db.execute(
        "SELECT t.*, c.name as cat_name FROM transactions t "
        "LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.description LIKE '%Taxes%'"
    ).fetchone()
    assert tax_txn["cat_name"] == "Payroll — Taxes"
    assert tax_txn["is_flagged"] == 0
